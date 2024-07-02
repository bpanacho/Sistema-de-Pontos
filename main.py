# Importando as bibliotecas
import cv2
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from PIL import Image as PILImage
import os
import sys
import tkinter as tk
from tkinter import messagebox

# Função para capturar a imagem do funcionário
def capture_image(employee_name):
    create_user_directory(employee_name)
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        cv2.imshow('Pressione "s" para capturar a foto', frame)
        if cv2.waitKey(1) & 0xFF == ord('s'):
            image_path = f"./users/{employee_name}/{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            cv2.imwrite(image_path, frame)
            break
    cap.release()
    cv2.destroyAllWindows()
    return image_path

# Função para criar diretório do usuário
def create_user_directory(employee_name):
    user_directory = f"./users/{employee_name}"
    if not os.path.exists(user_directory):
        os.makedirs(user_directory)

# Função para redimensionar a imagem
def resize_image(image_path, width, height):
    img = PILImage.open(image_path)
    img = img.resize((width, height), PILImage.LANCZOS)
    resized_path = f"./users/{os.path.basename(image_path)}"
    img.save(resized_path)
    return resized_path

# Função para registrar o ponto no arquivo Excel
def register_time(employee_name, image_path, tipo_registro):
    file_name = "registro_ponto.xlsx"
    password = "cleria123"

    # Verifica se o arquivo já existe
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()

    # Verifica se a aba do funcionário já existe
    if employee_name in wb.sheetnames:
        sheet = wb[employee_name]
    else:
        sheet = wb.create_sheet(employee_name)
        sheet.append(["Data/Hora", "Tipo", "Foto"])

    # Adiciona os registros de ponto
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Redimensiona a imagem para 150x150 pixels
    resized_image_path = resize_image(image_path, 150, 150)
    img = XLImage(resized_image_path)
    sheet.append([current_time, tipo_registro, ""])

    # Coloca a imagem na célula correspondente
    row = sheet.max_row
    img.anchor = f'C{row}'
    sheet.add_image(img)

    # Ajusta a largura da coluna para caber a data/hora e tipo de registro
    sheet.column_dimensions[get_column_letter(1)].width = 20
    sheet.column_dimensions[get_column_letter(2)].width = 10
    # Ajusta a altura da linha para acomodar a imagem
    sheet.row_dimensions[row].height = 100

    # Protege a planilha com a senha
    for ws in wb.worksheets:
        ws.protection.set_password(password)
        ws.protection.sheet = True

    # Salva o arquivo Excel
    wb.save(file_name)

# Função para ler a última ação registrada 
# Entrada ou saída
def read_last_action(employee_name):
    file_name = f"./users/{employee_name}_last_action.txt"
    if os.path.exists(file_name):
        with open(file_name, 'r') as file:
            return file.read().strip()
    return None

# Função para salvar a última ação registrada
def write_last_action(employee_name, action):
    file_name = f"./users/{employee_name}_last_action.txt"
    with open(file_name, 'w') as file:
        file.write(action)

def main(employee_name):
    while True:
        last_action = read_last_action(employee_name)

        if last_action == "Entrada":
            tipo_registro = "Saída"
        else:
            tipo_registro = "Entrada"

        image_path = capture_image(employee_name)
        register_time(employee_name, image_path, tipo_registro)
        write_last_action(employee_name, tipo_registro)

        # Mostra a mensagem de sucesso e pergunta sobre o próximo registro
        root = tk.Tk()
        root.withdraw()  # Esconde a janela principal
        result = messagebox.askyesno("Registro de Ponto", "\n Deseja registrar o próximo ponto?")
        root.destroy()  # Fecha a janela após a interação

        if not result:
            break

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Por favor, forneça o nome do funcionário.")
        sys.exit(1)
    
    employee_name = sys.argv[1]
    main(employee_name)
    
